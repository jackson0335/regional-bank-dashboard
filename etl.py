"""
ETL: databank_02_chigin_an.xlsx → 正規化Parquetファイル

banks.parquet  : bank_code, bank_name, hq_city
items.parquet  : item_id, item_name, category_large, category_mid, source, consolidation, unit
values.parquet : fiscal_year, bank_code, item_id, value
"""

from __future__ import annotations

import os
import re
import hashlib
import unicodedata
from typing import Optional, List, Dict, Tuple
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
EXCEL_PATH = Path(__file__).parent / "databank_02_chigin_an.xlsx"

YEAR_SHEET_RE = re.compile(r"^\d{6}$")


def _normalize(s: str) -> str:
    """全角英数を半角に統一してからstrip"""
    return unicodedata.normalize("NFKC", s).strip()


def _is_skip_bank(name: str) -> bool:
    """括弧付き銀行名（子会社等）やFG/HDはスキップ"""
    if not name:
        return True
    n = _normalize(name).replace(" ", "").replace("\u3000", "")
    if n in ("-", "－", ""):
        return True
    if n.startswith("(") or n.startswith("（"):
        return True
    if n == "地銀合計":
        return False
    skip_keywords = ["FG", "フィナンシャル", "ホールディングス", "HD", "FHD", "グループ"]
    for kw in skip_keywords:
        if kw in n:
            return True
    return False


def _add_bank_suffix(name: str) -> str:
    """銀行名に『銀行』が付いていなければ付ける。ただし特殊名は除く"""
    if not name:
        return name
    if name in ("地銀合計",):
        return name
    if name.endswith("銀行"):
        return name
    if any(name.endswith(s) for s in ("信金", "信用金庫", "信託")):
        return name
    return name + "銀行"


def _parse_value(v) -> float | None:
    """セル値を float | None に変換"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("-", "- ", "", "nan", "None", "－"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _extract_banks(ws, year_sheets: list[str]) -> pd.DataFrame:
    """Row 2=コード, Row 3=名前, Row 4=所在地 から銀行マスタを構築"""
    rows = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        rows[idx] = list(row)

    codes_row = rows[0]
    names_row = rows[1]
    cities_row = rows[2]

    banks = []
    bank_col_map = {}

    for col_idx in range(11, len(names_row)):
        name = names_row[col_idx]
        if name is None:
            continue
        name = _normalize(str(name))
        if not name:
            continue
        if _is_skip_bank(name):
            continue

        code_raw = codes_row[col_idx] if col_idx < len(codes_row) else None
        if code_raw is None:
            continue
        code = _normalize(str(code_raw))
        if code.startswith("("):
            continue

        city = ""
        if col_idx < len(cities_row) and cities_row[col_idx] is not None:
            city = _normalize(str(cities_row[col_idx]))

        display_name = _add_bank_suffix(name)

        bank_col_map[col_idx] = code
        banks.append({"bank_code": code, "bank_name": display_name, "hq_city": city})

    df = pd.DataFrame(banks).drop_duplicates(subset="bank_code")
    return df, bank_col_map


def _extract_items_and_values(ws, fiscal_year: str, bank_col_map: dict) -> tuple[list[dict], list[dict]]:
    """セクションヘッダとデータ行を解析"""
    items = []
    values = []
    seen_items = set()

    cat_large = ""
    cat_mid = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
        vals = list(row)

        if len(vals) < 12:
            continue

        col_j = vals[9] if len(vals) > 9 else None
        col_k = vals[10] if len(vals) > 10 else None

        if col_j is not None:
            sec = str(col_j).strip()
            if sec.startswith("●"):
                cat_large = sec
                cat_mid = ""
                continue
            elif sec.startswith("▼"):
                cat_mid = sec
                continue
            elif sec:
                if re.match(r"^[０-９\d]+[．.]", sec):
                    cat_mid = sec
                continue

        if col_k is None or str(col_k).strip() == "":
            continue

        item_name = str(col_k).strip()
        source = str(vals[4]).strip() if vals[4] is not None else ""
        consolidation = str(vals[5]).strip() if vals[5] is not None else ""
        unit = str(vals[6]).strip() if vals[6] is not None else ""

        item_key = f"{item_name}|{source}|{consolidation}|{unit}|{cat_large}|{cat_mid}"
        item_id = hashlib.md5(item_key.encode()).hexdigest()[:12]

        if item_id not in seen_items:
            seen_items.add(item_id)
            items.append({
                "item_id": item_id,
                "item_name": item_name,
                "category_large": cat_large,
                "category_mid": cat_mid,
                "source": source,
                "consolidation": consolidation,
                "unit": unit,
            })

        for col_idx, bank_code in bank_col_map.items():
            if col_idx < len(vals):
                v = _parse_value(vals[col_idx])
                if v is not None:
                    values.append({
                        "fiscal_year": fiscal_year,
                        "bank_code": bank_code,
                        "item_id": item_id,
                        "value": v,
                    })

    return items, values


def run_etl(force: bool = False) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """ETL実行。キャッシュがあればスキップ。"""
    banks_path = DATA_DIR / "banks.parquet"
    items_path = DATA_DIR / "items.parquet"
    values_path = DATA_DIR / "values.parquet"

    if not force and banks_path.exists() and items_path.exists() and values_path.exists():
        return (
            pd.read_parquet(banks_path),
            pd.read_parquet(items_path),
            pd.read_parquet(values_path),
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Loading workbook: {EXCEL_PATH}")
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    year_sheets = [s for s in wb.sheetnames if YEAR_SHEET_RE.match(s)]
    year_sheets.sort()
    print(f"Year sheets: {year_sheets}")

    first_ws = wb[year_sheets[0]]
    banks_df, bank_col_map = _extract_banks(first_ws, year_sheets)
    print(f"Banks extracted: {len(banks_df)}")
    print(f"Bank names: {banks_df['bank_name'].tolist()}")

    fg_check = banks_df[banks_df["bank_name"].str.contains("FG|フィナンシャル|ホールディングス|HD", na=False)]
    if len(fg_check) > 0:
        print(f"WARNING: FG banks still present: {fg_check['bank_name'].tolist()}")
    else:
        print("OK: No FG banks in data")

    all_items = []
    all_values = []

    for sheet_name in year_sheets:
        fiscal_year = sheet_name
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]

        _, sheet_bank_col_map = _extract_banks(ws, year_sheets)

        items, values = _extract_items_and_values(ws, fiscal_year, sheet_bank_col_map)
        all_items.extend(items)
        all_values.extend(values)
        print(f"  Items: {len(items)}, Values: {len(values)}")

    wb.close()

    items_df = pd.DataFrame(all_items).drop_duplicates(subset="item_id")
    values_df = pd.DataFrame(all_values)

    print(f"\nFinal: {len(banks_df)} banks, {len(items_df)} items, {len(values_df)} values")

    banks_df.to_parquet(banks_path, index=False)
    items_df.to_parquet(items_path, index=False)
    values_df.to_parquet(values_path, index=False)

    print("Parquet files saved to data/")
    return banks_df, items_df, values_df


if __name__ == "__main__":
    run_etl(force=True)
